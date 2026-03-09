"""
Æsethien Vocabulary Enrichment Pipeline
========================================
Sharon's workflow:
  1. Open data/aesethien-vocab.xlsx
  2. Add rows: Room | German Word  (columns A and B only)
  3. Save the file
  4. Run:  python tools/enrich.py
  5. Commit + push via GitHub Desktop

What this script does:
  - Reads the spreadsheet
  - Hits Wiktionary API for: article, plural, English meaning, IPA, audio URL
  - Downloads audio files to public/audio/
  - Hits Noun Project API for icons (needs API keys in .env)
  - Downloads icons to public/icons/
  - Writes enriched JSON to public/vocab/data/{room}.json
  - Updates the spreadsheet with all enriched data

Requirements:
  pip install -r tools/requirements.txt

API Keys (add to .env file in repo root):
  NOUN_PROJECT_KEY=your_key_here
  NOUN_PROJECT_SECRET=your_secret_here
"""

import os, re, sys, json, time, requests
from pathlib import Path
from openpyxl import load_workbook
from requests_oauthlib import OAuth1

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).parent.parent
XLSX       = ROOT / "data" / "aesethien-vocab.xlsx"
DATA_DIR   = ROOT / "public" / "vocab" / "data"
ICONS_DIR  = ROOT / "public" / "icons"
AUDIO_DIR  = ROOT / "public" / "audio"

DATA_DIR.mkdir(parents=True, exist_ok=True)
ICONS_DIR.mkdir(parents=True, exist_ok=True)
AUDIO_DIR.mkdir(parents=True, exist_ok=True)

# ── Load API keys from .env ───────────────────────────────────────────────────
def load_env():
    env_file = ROOT / ".env"
    env = {}
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

ENV = load_env()
NOUN_KEY    = ENV.get("NOUN_PROJECT_KEY", "")
NOUN_SECRET = ENV.get("NOUN_PROJECT_SECRET", "")

# ── Wiktionary ────────────────────────────────────────────────────────────────
ARTICLE_MAP = {"der": "der", "die": "die", "das": "das"}

def fetch_wiktionary(word):
    """Returns dict with article, plural, english, ipa, audio_filename"""
    result = {"article": "", "plural": "—", "english": "", "ipa": "", "audio_url": "", "audio_file": ""}
    try:
        url = f"https://en.wiktionary.org/w/api.php"
        params = {"action": "parse", "page": word, "prop": "wikitext", "format": "json"}
        r = requests.get(url, params=params, timeout=8)
        wikitext = r.json().get("parse", {}).get("wikitext", {}).get("*", "")

        # Article — look for {{de-noun|...}} or gender markers
        gender_match = re.search(r'\{\{de-noun\|([mfn])', wikitext)
        if gender_match:
            g = gender_match.group(1)
            result["article"] = {"m":"der","f":"die","n":"das"}.get(g,"")

        # Plural
        plural_match = re.search(r'pl[123]?=([^|}\n]+)', wikitext)
        if plural_match:
            result["plural"] = plural_match.group(1).strip()

        # English definition — first gloss
        gloss_match = re.search(r'# ([^\n{[]+)', wikitext)
        if gloss_match:
            result["english"] = gloss_match.group(1).strip().rstrip(".")

        # IPA
        ipa_match = re.search(r'\|([^|}\n]*[aeiouɐɪʊəɔæɛœøyʏ][^|}\n]*)\}\}', wikitext)
        if not ipa_match:
            ipa_match = re.search(r'IPA[^|]*\|([^|}]+)', wikitext)
        if ipa_match:
            result["ipa"] = ipa_match.group(1).strip()

        # Audio file — look for {{audio|de|De-{word}.ogg}}
        audio_match = re.search(r'audio\|de\|([^\|}\n]+\.ogg)', wikitext, re.IGNORECASE)
        if audio_match:
            audio_filename = audio_match.group(1).strip()
            result["audio_url"] = f"https://commons.wikimedia.org/wiki/Special:FilePath/{audio_filename}"
            result["audio_file"] = audio_filename

    except Exception as e:
        print(f"  ⚠ Wiktionary error for '{word}': {e}")
    return result

def download_audio(audio_url, audio_filename, word):
    """Download audio file to public/audio/"""
    if not audio_url:
        return ""
    dest = AUDIO_DIR / audio_filename
    if dest.exists():
        return audio_filename
    try:
        r = requests.get(audio_url, timeout=10, allow_redirects=True)
        if r.status_code == 200:
            dest.write_bytes(r.content)
            print(f"  ✓ Audio: {audio_filename}")
            return audio_filename
    except Exception as e:
        print(f"  ⚠ Audio download failed for '{word}': {e}")
    return ""

# ── Noun Project ──────────────────────────────────────────────────────────────
def fetch_noun_project_icon(english_term, word):
    """Download icon SVG from Noun Project. Returns filename or empty string."""
    if not NOUN_KEY or not NOUN_SECRET:
        print(f"  ℹ No Noun Project keys — skipping icon for '{word}'")
        return ""

    filename = f"{english_term.lower().replace(' ','_')}.svg"
    dest = ICONS_DIR / filename
    if dest.exists():
        return filename

    try:
        auth = OAuth1(NOUN_KEY, NOUN_SECRET)
        search_url = f"https://api.thenounproject.com/v2/icon?query={english_term}&limit=1&thumbnail_size=84"
        r = requests.get(search_url, auth=auth, timeout=8)
        data = r.json()
        icons = data.get("icons", [])
        if not icons:
            return ""

        # Download the SVG
        svg_url = icons[0].get("thumbnail_url") or icons[0].get("preview_url_84", "")
        if svg_url:
            svg_r = requests.get(svg_url, timeout=8)
            dest.write_bytes(svg_r.content)
            print(f"  ✓ Icon: {filename}")
            return filename
    except Exception as e:
        print(f"  ⚠ Noun Project error for '{word}': {e}")
    return ""

# ── Read spreadsheet ──────────────────────────────────────────────────────────
def read_spreadsheet():
    wb = load_workbook(XLSX)
    ws = wb["Vocabulary"]

    rooms = {}  # { room_name: [words] }

    for row in ws.iter_rows(min_row=4, values_only=True):
        room, word = row[0], row[1]
        if not room or not word:
            continue
        room = str(room).strip()
        word = str(word).strip()
        if room not in rooms:
            rooms[room] = []
        rooms[room].append(word)

    return wb, ws, rooms

# ── Write back to spreadsheet ─────────────────────────────────────────────────
def update_spreadsheet(wb, ws, enriched):
    """Write enriched data back into columns C–H"""
    row_idx = 4
    for row in ws.iter_rows(min_row=4):
        room_cell = row[0].value
        word_cell = row[1].value
        if not room_cell or not word_cell:
            row_idx += 1
            continue
        word = str(word_cell).strip()
        room = str(room_cell).strip()
        key = f"{room}::{word}"
        if key in enriched:
            e = enriched[key]
            ws.cell(row=row_idx, column=3, value=e.get("article",""))
            ws.cell(row=row_idx, column=4, value=e.get("plural",""))
            ws.cell(row=row_idx, column=5, value=e.get("english",""))
            ws.cell(row=row_idx, column=6, value=e.get("ipa",""))
            ws.cell(row=row_idx, column=7, value=e.get("audio_file",""))
            ws.cell(row=row_idx, column=8, value=e.get("icon_term",""))
        row_idx += 1
    wb.save(XLSX)
    print(f"\n✓ Spreadsheet updated: {XLSX}")

# ── Write JSON files ──────────────────────────────────────────────────────────
ROOM_META = {
    "Markthalle":        {"displayName": "Die Markthalle",        "description": "The market hall — food, quantities, prices", "session": 3},
    "Depot":             {"displayName": "Das Depot",             "description": "The depot — objects, tools, furniture",      "session": 9},
    "Kernarchiv":        {"displayName": "Das Kernarchiv",        "description": "The archive — documents, time, reading",     "session": 5},
    "Ankunftsbahnsteig": {"displayName": "Der Ankunftsbahnsteig", "description": "The arrival platform — greetings, numbers", "session": 1},
    "Haengegaerten":     {"displayName": "Die Hängegärten",       "description": "The hanging gardens — nature, directions",  "session": 7},
    "Ratssaal":          {"displayName": "Der Ratssaal",          "description": "The council chamber — opinions, debate",    "session": 11},
}

def write_json(room, words_data):
    meta = ROOM_META.get(room, {"displayName": room, "description": "", "session": 0})
    output = {
        "room": room,
        "displayName": meta["displayName"],
        "description": meta["description"],
        "session": meta["session"],
        "words": words_data
    }
    filename = room.lower().replace(" ", "_").replace("ä","ae").replace("ö","oe").replace("ü","ue") + ".json"
    dest = DATA_DIR / filename
    dest.write_text(json.dumps(output, ensure_ascii=False, indent=2))
    print(f"✓ JSON written: {dest.name} ({len(words_data)} words)")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("\n🌿 Æsethien Vocabulary Enrichment Pipeline")
    print("=" * 45)

    wb, ws, rooms = read_spreadsheet()
    print(f"📋 Found {sum(len(v) for v in rooms.values())} words across {len(rooms)} rooms\n")

    enriched = {}      # key: "Room::Word", value: enriched dict
    room_words = {}    # key: room, value: list of word dicts for JSON

    for room, words in rooms.items():
        print(f"\n📍 {room} ({len(words)} words)")
        room_words[room] = []

        for i, word in enumerate(words):
            print(f"  [{i+1}/{len(words)}] {word}")
            wiki = fetch_wiktionary(word)
            time.sleep(0.3)  # be polite to Wiktionary

            # Download audio
            audio_file = ""
            if wiki["audio_url"]:
                audio_file = download_audio(wiki["audio_url"], wiki["audio_file"], word)

            # Fetch icon
            icon_term = wiki["english"].split("/")[0].strip().split(",")[0].strip().lower()
            icon_file = fetch_noun_project_icon(icon_term, word) if icon_term else ""
            if NOUN_KEY:
                time.sleep(0.2)  # rate limit

            word_entry = {
                "id":       i + 1,
                "word":     word,
                "article":  wiki["article"],
                "plural":   wiki["plural"] or "—",
                "en":       wiki["english"],
                "ipa":      wiki["ipa"],
                "emoji":    "",          # kept empty — icons replace emoji in production
                "iconFile": icon_file,
                "audio":    audio_file,
            }

            room_words[room].append(word_entry)
            enriched[f"{room}::{word}"] = {**wiki, "icon_term": icon_term, "icon_file": icon_file}

        write_json(room, room_words[room])

    update_spreadsheet(wb, ws, enriched)
    print("\n✅ All done! Push to GitHub to go live.")
    print(f"   vocab index: https://sirnetizen.github.io/aesethien-world/vocab/")

if __name__ == "__main__":
    main()
